from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

from apps.imports.dtos import ExcelPreview
from apps.imports.exceptions import ExcelPreviewError


class ExcelPreviewService:
    XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
    XLS_EXTENSIONS = {".xls"}

    def get_preview(
        self,
        file_path: str,
        sheet_name: str | None = None,
        limit: int = 20,
    ) -> ExcelPreview:
        path = Path(file_path)
        extension = path.suffix.lower()

        if extension in self.XLSX_EXTENSIONS:
            return self._get_xlsx_preview(path, sheet_name, limit)

        if extension in self.XLS_EXTENSIONS:
            return self._get_xls_preview(path, sheet_name, limit)

        raise ExcelPreviewError(
            "Unsupported file format. Only .xlsx, .xlsm and .xls are supported."
        )

    def _get_xlsx_preview(
        self,
        path: Path,
        sheet_name: str | None,
        limit: int,
    ) -> ExcelPreview:
        try:
            workbook = openpyxl.load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ExcelPreviewError(f"Cannot read Excel file: {exc}") from exc

        try:
            sheet_names = workbook.sheetnames

            if not sheet_names:
                raise ExcelPreviewError("Excel file does not contain sheets.")

            if sheet_name:
                if sheet_name not in sheet_names:
                    raise ExcelPreviewError(f"Sheet '{sheet_name}' was not found.")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            rows: list[list[Any]] = []

            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= limit:
                    break

                rows.append([self._normalize_cell(value) for value in row])

            return ExcelPreview(
                sheet_name=sheet.title,
                sheet_names=sheet_names,
                rows=rows,
            )
        finally:
            workbook.close()

    def _get_xls_preview(
        self,
        path: Path,
        sheet_name: str | None,
        limit: int,
    ) -> ExcelPreview:
        try:
            workbook = xlrd.open_workbook(path)
        except Exception as exc:
            raise ExcelPreviewError(f"Cannot read Excel file: {exc}") from exc

        sheet_names = workbook.sheet_names()

        if not sheet_names:
            raise ExcelPreviewError("Excel file does not contain sheets.")

        if sheet_name:
            if sheet_name not in sheet_names:
                raise ExcelPreviewError(f"Sheet '{sheet_name}' was not found.")
            sheet = workbook.sheet_by_name(sheet_name)
        else:
            sheet = workbook.sheet_by_index(0)

        rows: list[list[Any]] = []
        rows_count = min(sheet.nrows, limit)

        for row_index in range(rows_count):
            row = sheet.row_values(row_index)
            rows.append([self._normalize_cell(value) for value in row])

        return ExcelPreview(
            sheet_name=sheet.name,
            sheet_names=sheet_names,
            rows=rows,
        )

    def _normalize_cell(self, value: Any) -> Any:
        if value is None:
            return None

        if isinstance(value, datetime | date | time):
            return value.isoformat()

        if isinstance(value, Decimal):
            return str(value)

        return value

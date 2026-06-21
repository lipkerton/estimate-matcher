from collections.abc import Iterator
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

from apps.imports.dtos import ExcelRow
from apps.imports.exceptions import ExcelReadError


class ExcelRowReader:
    XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
    XLS_EXTENSIONS = {".xls"}

    def get_total_rows(self, file_path: str, sheet_name: str | None = None) -> int:
        path = Path(file_path)
        extension = path.suffix.lower()

        if extension in self.XLSX_EXTENSIONS:
            return self._get_xlsx_total_rows(path, sheet_name)

        if extension in self.XLS_EXTENSIONS:
            return self._get_xls_total_rows(path, sheet_name)

        raise ExcelReadError(
            "Unsupported file format. Only .xlsx, .xlsm and .xls are supported."
        )

    def iter_rows(
        self,
        file_path: str,
        sheet_name: str | None = None,
        start_row: int = 0,
    ) -> Iterator[ExcelRow]:
        path = Path(file_path)
        extension = path.suffix.lower()

        if extension in self.XLSX_EXTENSIONS:
            yield from self._iter_xlsx_rows(path, sheet_name, start_row)
            return

        if extension in self.XLS_EXTENSIONS:
            yield from self._iter_xls_rows(path, sheet_name, start_row)
            return

        raise ExcelReadError(
            "Unsupported file format. Only .xlsx, .xlsm and .xls are supported."
        )

    def _get_xlsx_total_rows(self, path: Path, sheet_name: str | None) -> int:
        try:
            workbook = openpyxl.load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ExcelReadError(f"Cannot read Excel file: {exc}") from exc

        try:
            sheet = self._get_xlsx_sheet(workbook, sheet_name)
            return sheet.max_row
        finally:
            workbook.close()

    def _iter_xlsx_rows(
        self,
        path: Path,
        sheet_name: str | None,
        start_row: int,
    ) -> Iterator[ExcelRow]:
        try:
            workbook = openpyxl.load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ExcelReadError(f"Cannot read Excel file: {exc}") from exc

        try:
            sheet = self._get_xlsx_sheet(workbook, sheet_name)

            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index < start_row:
                    continue

                yield ExcelRow(
                    row_number=row_index + 1,
                    values=[self._normalize_cell(value) for value in row],
                )
        finally:
            workbook.close()

    def _get_xlsx_sheet(self, workbook, sheet_name: str | None):
        sheet_names = workbook.sheetnames

        if not sheet_names:
            raise ExcelReadError("Excel file does not contain sheets.")

        if sheet_name:
            if sheet_name not in sheet_names:
                raise ExcelReadError(f"Sheet '{sheet_name}' was not found.")
            return workbook[sheet_name]

        return workbook.active

    def _get_xls_total_rows(self, path: Path, sheet_name: str | None) -> int:
        try:
            workbook = xlrd.open_workbook(path)
        except Exception as exc:
            raise ExcelReadError(f"Cannot read Excel file: {exc}") from exc

        sheet = self._get_xls_sheet(workbook, sheet_name)
        return sheet.nrows

    def _iter_xls_rows(
        self,
        path: Path,
        sheet_name: str | None,
        start_row: int,
    ) -> Iterator[ExcelRow]:
        try:
            workbook = xlrd.open_workbook(path)
        except Exception as exc:
            raise ExcelReadError(f"Cannot read Excel file: {exc}") from exc

        sheet = self._get_xls_sheet(workbook, sheet_name)

        for row_index in range(start_row, sheet.nrows):
            yield ExcelRow(
                row_number=row_index + 1,
                values=[
                    self._normalize_cell(value)
                    for value in sheet.row_values(row_index)
                ],
            )

    def _get_xls_sheet(self, workbook, sheet_name: str | None):
        sheet_names = workbook.sheet_names()

        if not sheet_names:
            raise ExcelReadError("Excel file does not contain sheets.")

        if sheet_name:
            if sheet_name not in sheet_names:
                raise ExcelReadError(f"Sheet '{sheet_name}' was not found.")
            return workbook.sheet_by_name(sheet_name)

        return workbook.sheet_by_index(0)

    def _normalize_cell(self, value: Any) -> Any:
        if value is None:
            return None

        if isinstance(value, datetime | date | time):
            return value.isoformat()

        if isinstance(value, Decimal):
            return str(value)

        return value
